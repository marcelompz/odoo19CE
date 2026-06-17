# -*- coding: utf-8 -*-

import base64
import io
from odoo import models, fields, api, _
from odoo.exceptions import UserError

try:
    import openpyxl
except ImportError:
    openpyxl = None


class ProductMassImportWizard(models.TransientModel):
    _name = 'product.mass.import.wizard'
    _description = 'Wizard de Importación Masiva de Productos desde Excel'

    file_data = fields.Binary(string='Archivo Excel (.xlsx)', required=True)
    filename = fields.Char(string='Nombre del Archivo')
    location_id = fields.Many2one(
        'stock.location',
        string='Ubicación de Inventario',
        required=True,
        domain=[('usage', '=', 'internal')],
        help="Ubicación física donde se cargará el stock inicial por defecto."
    )
    state = fields.Selection([
        ('draft', 'Borrador'),
        ('preview', 'Vista Previa'),
        ('done', 'Completado'),
    ], default='draft', string='Estado')
    product_count = fields.Integer(string='Cantidad de Productos', compute='_compute_product_count')
    preview_ids = fields.One2many('product.mass.import.preview', 'wizard_id', string='Vista Previa')

    @api.model
    def default_get(self, fields_list):
        res = super(ProductMassImportWizard, self).default_get(fields_list)
        warehouse = self.env['stock.warehouse'].search([], limit=1)
        if warehouse:
            res['location_id'] = warehouse.lot_stock_id.id
        return res

    def _compute_product_count(self):
        for wizard in self:
            wizard.product_count = len(wizard.preview_ids)

    def action_download_template(self):
        """Download Excel template for product import"""
        if not openpyxl:
            raise UserError(_("La librería 'openpyxl' es necesaria. Ejecute: pip install openpyxl"))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Plantilla Productos'

        # Headers
        headers = [
            'Referencia Interna',
            'Nombre del Producto',
            'Descripción para PdV',
            'Código de Barras',
            'Disponible en PdV',
            'Categoría de Producto',
            'Categoría de PdV',
            'Precio de Venta',
            'Precio de Costo',
            'Cantidad a la Mano',
            'Tipo de Producto',
            'Trazabilidad'
        ]
        ws.append(headers)

        # Example row
        example = [
            'PROD-001',
            'Producto Ejemplo',
            'Descripción para mostrar en punto de venta',
            '7701234567890',
            'VERDADERO',
            'Electrodomésticos',
            'Electrodomésticos',
            100000.00,
            75000.00,
            50,
            'Almacenable',
            'Ninguno'
        ]
        ws.append(example)

        # Column widths
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 20

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        file_content = base64.b64encode(output.read())

        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/?model=product.mass.import.wizard&id={self.id}&field=file_data&download=true&filename=plantilla_productos.xlsx',
            'target': 'new',
        }

    def action_parse_excel(self):
        """Parse Excel file and show preview"""
        if not openpyxl:
            raise UserError(_("La librería 'openpyxl' es necesaria. Ejecute: pip install openpyxl"))

        self.ensure_one()
        self.preview_ids.unlink()

        try:
            data = base64.b64decode(self.file_data)
            wb = openpyxl.load_workbook(filename=io.BytesIO(data), data_only=True)
            sheet = wb.active

            rows = list(sheet.iter_rows(min_row=2, values_only=True))
            preview_data = []
            errors = []

            for idx, row in enumerate(rows, start=2):
                if not row or not row[0]:
                    continue

                error_msgs = []
                default_code = str(row[0]).strip() if row[0] else False
                name = str(row[1]).strip() if row[1] else False
                pos_description = str(row[2]).strip() if row[2] else False
                barcode = str(row[3]).strip() if row[3] else False
                available_in_pos = str(row[4]).upper() in ['VERDADERO', 'TRUE', '1', 'SI'] if row[4] is not None else True
                categ_name = str(row[5]).strip() if row[5] else False
                pos_categ_name = str(row[6]).strip() if row[6] else False
                list_price = float(row[7]) if row[7] else 0.0
                standard_price = float(row[8]) if row[8] else 0.0
                qty_on_hand = float(row[9]) if row[9] else 0.0

                # Product type
                product_type = 'product'
                if row[10]:
                    type_val = str(row[10]).lower()
                    if type_val in ['consumible', 'consu']:
                        product_type = 'consu'
                    elif type_val in ['servicio', 'service']:
                        product_type = 'service'

                # Tracking
                tracking = 'none'
                if row[11]:
                    track_val = str(row[11]).lower()
                    if 'lote' in track_val:
                        tracking = 'lot'
                    elif 'serie' in track_val:
                        tracking = 'serial'

                # Validate barcode uniqueness
                if barcode:
                    existing = self.env['product.product'].search([('barcode', '=', barcode)], limit=1)
                    if existing:
                        error_msgs.append(f"Código de barras duplicado: {existing.name}")

                # Validate required fields
                if not default_code:
                    error_msgs.append("Referencia interna requerida")

                if not name:
                    error_msgs.append("Nombre del producto requerido")

                # Validate numeric fields (only if provided)
                if list_price is not None and list_price < 0:
                    error_msgs.append("Precio de venta no puede ser negativo")

                if standard_price is not None and standard_price < 0:
                    error_msgs.append("Precio de costo no puede ser negativo")

                if qty_on_hand is not None and qty_on_hand < 0:
                    error_msgs.append("Cantidad no puede ser negativa")

                error_str = ', '.join(error_msgs) if error_msgs else ''

                preview_data.append((0, 0, {
                    'row_number': idx,
                    'default_code': default_code or '',
                    'name': name or '',
                    'pos_description': pos_description or '',
                    'barcode': barcode or '',
                    'available_in_pos': available_in_pos,
                    'categ_name': categ_name or '',
                    'pos_categ_name': pos_categ_name or '',
                    'list_price': list_price,
                    'standard_price': standard_price,
                    'qty_on_hand': qty_on_hand,
                    'product_type': product_type,
                    'tracking': tracking,
                    'error_message': error_str,
                    'is_valid': len(error_msgs) == 0,
                }))

            self.write({
                'state': 'preview',
                'preview_ids': preview_data,
            })

            valid_count = sum(1 for p in preview_data if p[2]['is_valid'])
            invalid_count = len(preview_data) - valid_count

            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Análisis Completado'),
                    'message': _('Productos válidos: %d, Con errores: %d') % (valid_count, invalid_count),
                    'type': 'success' if invalid_count == 0 else 'warning',
                    'sticky': invalid_count > 0,
                }
            }

        except Exception as e:
            raise UserError(_("Error al procesar el archivo Excel: %s") % str(e))

    def action_confirm_import(self):
        """Confirm import and create products"""
        self.ensure_one()

        if not self.preview_ids:
            raise UserError(_("No hay productos para importar"))

        valid_products = self.preview_ids.filtered(lambda p: p.is_valid)
        invalid_products = self.preview_ids.filtered(lambda p: not p.is_valid)

        if not valid_products:
            raise UserError(_("No hay productos válidos para importar. Corrija los errores primero."))

        created_products = []
        products_to_quant = []

        for preview in valid_products:
            # Get or create product category
            categ_id = self.env.ref('product.product_category_all').id
            if preview.categ_name:
                category = self.env['product.category'].search([('name', '=', preview.categ_name)], limit=1)
                if not category:
                    category = self.env['product.category'].create({'name': preview.categ_name})
                categ_id = category.id

            # Get or create POS category
            pos_categ_id = False
            if preview.pos_categ_name and 'pos.category' in self.env:
                pos_category = self.env['pos.category'].search([('name', '=', preview.pos_categ_name)], limit=1)
                if not pos_category:
                    pos_category = self.env['pos.category'].create({'name': preview.pos_categ_name})
                pos_categ_id = pos_category.id

            product_vals = {
                'name': preview.name,
                'default_code': preview.default_code,
                'barcode': preview.barcode or False,
                'list_price': preview.list_price,
                'standard_price': preview.standard_price,
                'type': preview.product_type,
                'categ_id': categ_id,
                'tracking': preview.tracking,
                'available_in_pos': preview.available_in_pos,
            }

            if preview.pos_description:
                product_vals['description_sale'] = preview.pos_description

            if pos_categ_id:
                product_vals['pos_categ_id'] = pos_categ_id

            product = self.env['product.product'].create(product_vals)
            created_products.append(product)

            if preview.qty_on_hand > 0:
                products_to_quant.append((product, preview.qty_on_hand))

        # Apply inventory quantities
        if products_to_quant:
            for product, qty in products_to_quant:
                self.env['stock.quant'].with_context(inventory_mode=True).create({
                    'product_id': product.id,
                    'location_id': self.location_id.id,
                    'inventory_quantity': qty,
                }).action_apply_inventory()

        self.write({'state': 'done'})

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Proceso Completado'),
                'message': _('Se crearon %d productos exitosamente.') % len(created_products),
                'type': 'success',
                'next': {'type': 'ir.actions.act_window_close'},
            }
        }


class ProductMassImportPreview(models.TransientModel):
    _name = 'product.mass.import.preview'
    _description = 'Vista Previa de Importación de Productos'
    _order = 'row_number'

    wizard_id = fields.Many2one('product.mass.import.wizard', string='Wizard', ondelete='cascade')
    row_number = fields.Integer(string='Fila')
    default_code = fields.Char(string='Referencia Interna')
    name = fields.Char(string='Nombre del Producto')
    pos_description = fields.Char(string='Descripción para PdV')
    barcode = fields.Char(string='Código de Barras')
    available_in_pos = fields.Boolean(string='Disponible en PdV')
    categ_name = fields.Char(string='Categoría de Producto')
    pos_categ_name = fields.Char(string='Categoría de PdV')
    list_price = fields.Float(string='Precio de Venta')
    standard_price = fields.Float(string='Precio de Costo')
    qty_on_hand = fields.Float(string='Cantidad a la Mano')
    product_type = fields.Selection([
        ('product', 'Almacenable'),
        ('consu', 'Consumible'),
        ('service', 'Servicio'),
    ], string='Tipo de Producto')
    tracking = fields.Selection([
        ('none', 'Ninguno'),
        ('lot', 'Por Lote'),
        ('serial', 'Por Número de Serie'),
    ], string='Trazabilidad')
    error_message = fields.Text(string='Errores')
    is_valid = fields.Boolean(string='Válido')
